# -*- coding: utf-8 -*-
"""工具函数：地址解析、高德地理编码、Excel 导入导出"""
import io
import json
import re
import secrets
import string
from typing import Optional, Tuple, List, Dict

import requests
from openpyxl import Workbook, load_workbook

# ===== 城市/区识别 =====
# 仅穷举珠三角常见城市；未匹配时回退至全国通用正则
PRD_CITIES = ["广州市", "深圳市", "东莞市", "惠州市", "江门市", "珠海市", "中山市", "佛山市", "肇庆市"]


def parse_city_district(address: str) -> Tuple[str, str, str]:
    """从地址中提取 (省, 市, 区/县)"""
    if not address:
        return "广东省", "", ""
    province = "广东省"
    m_prov = re.search(r"([\u4e00-\u9fa5]{2,6}省|北京市|上海市|天津市|重庆市|香港|澳门)", address)
    if m_prov:
        province = m_prov.group(1)

    city = ""
    for c in PRD_CITIES:
        if c in address:
            city = c
            break
    if not city:
        m_city = re.search(r"([\u4e00-\u9fa5]{2,8}市)", address)
        if m_city:
            city = m_city.group(1)

    district = ""
    # 在"市"之后再去识别区/县/镇/街道，避免吃到前面的"省/市"字
    tail = address
    if city and city in address:
        tail = address.split(city, 1)[1]
    # 兼容源数据异常：tail 开头若仍带其他地级市名（如"鹤山市石排镇"误前缀），剥离掉
    # 只对已知地级市做剥离，避免把"石碣镇城市…"中的"石碣镇城市"误当成市名
    KNOWN_EXTRA = ("鹤山市", "佛山市", "广州市", "深圳市", "东莞市", "惠州市", "江门市", "珠海市", "中山市", "肇庆市")
    for ec in KNOWN_EXTRA:
        if tail.startswith(ec):
            tail = tail[len(ec):]
            break
    # 优先匹配紧跟在市后面的 "X区"/"X县"
    m_dist = re.match(r"^([\u4e00-\u9fa5]{2,4}?(?:区|县))", tail)
    if m_dist:
        district = m_dist.group(1)
    else:
        # 东莞/中山等无区设置的地级市，取紧跟其后的 镇/街道
        m_town = re.match(r"^([\u4e00-\u9fa5]{2,5}?(?:镇|街道))", tail)
        if m_town:
            district = m_town.group(1)
        else:
            # 东莞功能区：松山湖、滨海湾、水乡新城等
            m_func = re.match(r"^(松山湖|滨海湾|水乡|银瓶合作创新区|粤海银瓶|东部工业园)", tail)
            if m_func:
                district = m_func.group(1)
    return province, city, district


# ===== 高德地理编码（地址 -> 经纬度） =====
def amap_geocode(address: str, city: str, key: str, name: str = "") -> Optional[Tuple[float, float]]:
    """调用高德 Web 服务 API 进行地理编码。
    优先使用结构化地理编码 /v3/geocode/geo；失败时回退到 PoI 搜索 /v3/place/text 兼容模糊地址。
    """
    if not key or not address:
        return None
    # 1) 标准地理编码
    try:
        r = requests.get(
            "https://restapi.amap.com/v3/geocode/geo",
            params={"address": address, "city": city or "", "key": key, "output": "JSON"},
            timeout=8,
        )
        data = r.json()
        if data.get("status") == "1" and data.get("geocodes"):
            loc = data["geocodes"][0].get("location", "")
            if "," in loc:
                lng, lat = loc.split(",")
                return float(lng), float(lat)
    except Exception:
        pass

    # 2) 回退：PoI 关键字搜索（去掉括号、用驿站名 + 地址精简关键词）
    try:
        # 把地址里的括号内容抽出（常常是真实 PoI 名）
        kw_parts = []
        if name:
            kw_parts.append(name)
        # 提取地址里的括号内容（例如"泊寓(千灯湖店)"）
        bracket = re.findall(r"[（(]([^（）()]+)[）)]", address)
        kw_parts.extend(bracket)
        # 取地址主体最末一段
        tail_addr = re.sub(r"[（(].*?[)）]", "", address)
        kw_parts.append(tail_addr[-30:])
        keyword = " ".join(dict.fromkeys([p for p in kw_parts if p]).keys())[:80]

        r = requests.get(
            "https://restapi.amap.com/v3/place/text",
            params={
                "keywords": keyword,
                "city": city or "",
                "citylimit": "true",
                "key": key,
                "offset": 1,
                "page": 1,
                "extensions": "base",
                "output": "JSON",
            },
            timeout=8,
        )
        data = r.json()
        if data.get("status") == "1" and data.get("pois"):
            loc = data["pois"][0].get("location", "")
            if "," in loc:
                lng, lat = loc.split(",")
                return float(lng), float(lat)
    except Exception:
        return None
    return None


# ===== 驿站码 =====
def generate_code(length: int = 8) -> str:
    alphabet = string.ascii_uppercase + string.digits
    # 去掉易混淆字符
    alphabet = alphabet.replace("0", "").replace("O", "").replace("1", "").replace("I", "")
    return "".join(secrets.choice(alphabet) for _ in range(length))


# ===== Excel 导入：兼容"工作地图"导出格式 =====
TEMPLATE_COLUMNS = ["名称", "城市", "区域", "地址", "申请入住指导", "备注", "文件夹", "位置编码"]


def read_stations_from_excel(file_stream) -> List[Dict]:
    """读取 Excel，兼容两种模板：
    1) 标准模板：名称/城市/区域/地址/申请入住指导/备注/文件夹/位置编码
    2) 工作地图导出："标记位置"sheet，含 名称/地址/描述/文件夹/位置编码 等
    """
    wb = load_workbook(file_stream, data_only=True)
    # 优先寻找标准模板的 Sheet1 或 "驿站信息"
    target_sheet = None
    for name in ["驿站信息", "Sheet1", "标记位置"]:
        if name in wb.sheetnames:
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb[wb.sheetnames[0]]

    rows = list(target_sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    records = []
    for raw in rows[1:]:
        item = {h: (raw[i] if i < len(raw) else None) for i, h in enumerate(header)}
        # 字段归一化
        name = item.get("名称") or item.get("驿站名称")
        if not name:
            continue
        address = item.get("地址") or ""
        guide = item.get("申请入住指导") or item.get("描述") or ""
        remark = item.get("备注") or item.get("不通过原因") or ""
        city = item.get("城市") or ""
        district = item.get("区域") or item.get("区") or ""
        folder = item.get("文件夹") or ""
        loc_code = item.get("位置编码(导出的内容若需再次导入，请不要修改此列)") or item.get("位置编码") or ""
        records.append({
            "name": str(name).strip(),
            "city": str(city).strip(),
            "district": str(district).strip(),
            "address": str(address).strip(),
            "guide_html": str(guide) if guide else "",
            "remark": str(remark) if remark else "",
            "folder": str(folder).strip() if folder else "",
            "location_code": str(loc_code).strip() if loc_code else "",
        })
    return records


def export_stations_to_excel(stations) -> bytes:
    """按标准模板导出"""
    wb = Workbook()
    ws = wb.active
    ws.title = "驿站信息"
    headers = ["名称", "城市", "区域", "地址", "经度", "纬度", "申请入住指导", "备注", "文件夹", "位置编码"]
    ws.append(headers)
    for s in stations:
        ws.append([
            s.name, s.city or "", s.district or "", s.address or "",
            s.lng or "", s.lat or "",
            _strip_html(s.guide_html or ""),
            s.remark or "",
            s.folder or "",
            s.location_code or "",
        ])
    # 列宽
    for col, w in zip("ABCDEFGHIJ", [28, 12, 12, 50, 12, 12, 60, 30, 18, 24]):
        ws.column_dimensions[col].width = w
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_import_template() -> bytes:
    """生成标准导入模板，供后台下载"""
    wb = Workbook()
    ws = wb.active
    ws.title = "驿站信息"
    ws.append(TEMPLATE_COLUMNS)
    ws.append([
        "示例：禅城青年驿站（泊寓·禅城中心旗舰店）",
        "佛山市", "禅城区",
        "广东省佛山市禅城区张槎街道马鞍街8号",
        "申请方式：登录XX小程序提交申请；联系人：张老师 13800000000",
        "免费入住7天，需提供毕业证/录用证明",
        "佛山青年人才驿站", "",
    ])
    for col, w in zip("ABCDEFGH", [30, 12, 12, 50, 60, 30, 22, 24]):
        ws.column_dimensions[col].width = w
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _strip_html(html: str) -> str:
    return re.sub(r"<[^>]+>", "", html or "").strip()


def safe_filename(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|]", "_", name)
    return name[:80] or "file"


# ===== 城市政策入口种子数据 =====
DEFAULT_CITY_POLICIES = [
    # 广州
    ("广州市", "人才认定", "🏅", "广州市青年人才认定申报", "申请青年人才驿站及各类补贴的前置认定", "https://hrss.gz.gov.cn/", 10),
    ("广州市", "生活补贴", "💰", "广州市青年人才生活津贴", "本科 1.5 万 / 硕士 2.5 万 / 博士 3 万一次性", "https://www.gz.gov.cn/zwfw/", 20),
    ("广州市", "租房补贴", "🏠", "广州青年人才租房补贴申报", "穗好办 App / 广州市住建局官网", "https://zfcj.gz.gov.cn/", 30),
    ("广州市", "落户", "🆔", "广州人才入户办理指引", "学历入户 / 技术入户 / 高层次人才入户", "https://hrss.gz.gov.cn/zwfw/grbsfw/rcrh/", 40),
    # 深圳
    ("深圳市", "人才认定", "🏅", "深圳市新引进人才租房和生活补贴", "本科 1.5 万 / 硕士 2.5 万 / 博士 3 万", "https://hrss.sz.gov.cn/", 10),
    ("深圳市", "生活补贴", "💰", "深圳青年驿站及补贴申请总入口", "深圳人社局青年人才服务专栏", "https://hrss.sz.gov.cn/szsi/qnyz/", 20),
    ("深圳市", "落户", "🆔", "深圳人才引进秒批", "应届生秒批 / 在职人才引进", "https://hrss.sz.gov.cn/szsi/", 30),
    # 东莞
    ("东莞市", "人才认定", "🏅", "东莞特色人才认定", "对应青年驿站及各项扶持的前置认定", "http://rsj.dg.gov.cn/", 10),
    ("东莞市", "生活补贴", "💰", "东莞青年人才驿站申请入口", "东莞人才网 / 莞家政务", "http://www.dg114.gov.cn/", 20),
    ("东莞市", "租房补贴", "🏠", "东莞市新引进人才综合补贴", "学历对应 1-30 万综合补贴", "http://rsj.dg.gov.cn/", 30),
    # 佛山
    ("佛山市", "人才认定", "🏅", "佛山市优秀人才认定", "I类至D类人才认定与申报", "http://hrss.foshan.gov.cn/", 10),
    ("佛山市", "生活补贴", "💰", "佛山青年人才安居补贴", "购房/租房补贴 / 生活补贴", "http://hrss.foshan.gov.cn/", 20),
    ("佛山市", "驿站", "🏨", "佛山青年人才驿站申请", "禅城/南海/顺德/高明 全市驿站统一入口", "http://hrss.foshan.gov.cn/", 30),
    # 珠海
    ("珠海市", "人才认定", "🏅", "珠海人才分类认定", "对应青年驿站、租房补贴、购房补贴", "http://zhrlzy.zhuhai.gov.cn/", 10),
    ("珠海市", "生活补贴", "💰", "珠海青年人才生活补助", "硕士 6 万 / 博士 10 万一次性", "http://zhrlzy.zhuhai.gov.cn/", 20),
    ("珠海市", "租房补贴", "🏠", "珠海高校毕业生租房补贴", "本科 7200/年，最长 3 年", "http://zhrlzy.zhuhai.gov.cn/", 30),
    # 中山
    ("中山市", "人才认定", "🏅", "中山英才计划申报", "对应青年驿站及各项扶持", "http://hrss.zs.gov.cn/", 10),
    ("中山市", "生活补贴", "💰", "中山青年人才驿站及补贴", "中山人社官方公告", "http://hrss.zs.gov.cn/", 20),
    # 惠州
    ("惠州市", "人才认定", "🏅", "惠州市人才分类目录", "对应人才驿站及购租补贴", "http://hrss.huizhou.gov.cn/", 10),
    ("惠州市", "生活补贴", "💰", "惠州青年人才入驻补贴", "惠州人社官方公告", "http://hrss.huizhou.gov.cn/", 20),
]


def seed_city_policies():
    """首次启动时填充政策入口种子数据"""
    from models import CityPolicy, db
    for city, cat, icon, title, desc, url, sort in DEFAULT_CITY_POLICIES:
        if not CityPolicy.query.filter_by(city=city, title=title).first():
            db.session.add(CityPolicy(
                city=city, category=cat, icon=icon, title=title,
                description=desc, url=url, sort=sort, enabled=True,
            ))
    db.session.commit()
    print(f"[SEED] 已植入 {len(DEFAULT_CITY_POLICIES)} 条城市政策入口")


# ===== ⑨ AI 驿站助手 =====
# 多服务商 + 知识库 RAG + 联网搜索 + 流式输出

import json as _json_

def _build_kb_context(question: str, use_knowledge: bool = True):
    """构建上下文：驿站 + 政策 + 通用知识库（FAQ + 多模态 RAG）
    返回: (上下文文本, citations 列表)
    """
    from models import Station, CityPolicy, KnowledgeEntry
    ctx_parts = []
    citations = []

    # 1) 提取问题里提到的城市
    cities = ["广州", "深圳", "东莞", "佛山", "珠海", "中山", "惠州", "江门", "肇庆"]
    hit_cities = [c for c in cities if c in question]

    # 2) 抽取相关驿站
    sq = Station.query
    if hit_cities:
        from sqlalchemy import or_ as _or
        sq = sq.filter(_or(*[Station.city.like(f"%{c}%") for c in hit_cities]))
    stations = sq.limit(15).all()
    if stations:
        lines = ["【相关驿站】"]
        for s in stations:
            piece = f"- {s.name}（{s.city}{s.district or ''}）地址：{s.address or '—'}"
            if s.contact_phone:
                piece += f"，电话 {s.contact_phone}"
            if s.contact_name:
                piece += f"（{s.contact_name}）"
            if s.free_days:
                piece += f"，免费 {s.free_days} 天"
            if s.apply_url:
                piece += f"，申请链接：{s.apply_url}"
            lines.append(piece)
        ctx_parts.append("\n".join(lines))

    # 3) 政策入口
    pq = CityPolicy.query.filter_by(enabled=True)
    if hit_cities:
        from sqlalchemy import or_ as _or
        pq = pq.filter(_or(*[CityPolicy.city.like(f"%{c}%") for c in hit_cities]))
    policies = pq.order_by(CityPolicy.sort).limit(12).all()
    if policies:
        lines = ["【相关政策入口】"]
        for p in policies:
            lines.append(f"- {p.city} {p.category} | {p.title}：{p.description or ''} 链接：{p.url}")
        ctx_parts.append("\n".join(lines))

    # 4) 多模态知识库（BM25 检索）
    if use_knowledge:
        try:
            from kb_ingest import build_rag_context_v2
            rag_text, rag_cites = build_rag_context_v2(question, top_k=5)
            if rag_text:
                ctx_parts.append(rag_text)
                citations.extend(rag_cites)
        except Exception as e:
            ctx_parts.append(f"（RAG 检索异常：{e}）")

        # 5) 兼容老 FAQ 表（KnowledgeEntry）
        kq = KnowledgeEntry.query.filter_by(enabled=True)
        kws = [k for k in re.findall(r"[\u4e00-\u9fa5A-Za-z0-9]{2,6}", question) if k]
        from sqlalchemy import or_ as _or
        if kws:
            conds = []
            for k in kws[:6]:
                conds.append(KnowledgeEntry.question.like(f"%{k}%"))
                conds.append(KnowledgeEntry.answer.like(f"%{k}%"))
                conds.append(KnowledgeEntry.keywords.like(f"%{k}%"))
            entries = kq.filter(_or(*conds)).limit(5).all()
        else:
            entries = kq.order_by(KnowledgeEntry.sort).limit(3).all()
        if entries:
            lines = ["【常见问答（FAQ）】"]
            for e in entries:
                lines.append(f"Q: {e.question}\nA: {e.answer[:300]}")
            ctx_parts.append("\n".join(lines))

    ctx_text = "\n\n".join(ctx_parts) if ctx_parts else "（暂无匹配数据）"
    return ctx_text, citations


def web_search(query: str, max_results: int = 5) -> str:
    """简易联网搜索：用 DuckDuckGo HTML 端点（无需 Key，国内可访问，速度快）。
    返回拼接好的文本片段，注入到 prompt 中。
    """
    try:
        from urllib.parse import quote
        # DuckDuckGo HTML
        r = requests.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query},
            headers={"User-Agent": "Mozilla/5.0 (compatible; YouthStation/1.0)"},
            timeout=8,
        )
        if r.status_code != 200:
            return ""
        # 简单提取 <a class="result__a" ...>title</a> + <a class="result__snippet">snippet
        results = []
        # 提取标题与链接
        for m in re.finditer(r'<a[^>]*class="[^"]*result__a[^"]*"[^>]*href="([^"]+)"[^>]*>(.*?)</a>', r.text, re.S):
            url, title = m.group(1), re.sub(r"<[^>]+>", "", m.group(2)).strip()
            if title and url:
                results.append({"title": title, "url": url, "snippet": ""})
            if len(results) >= max_results:
                break
        # 摘要
        snippets = []
        for m in re.finditer(r'<a[^>]*class="[^"]*result__snippet[^"]*"[^>]*>(.*?)</a>', r.text, re.S):
            snippets.append(re.sub(r"<[^>]+>", "", m.group(1)).strip())
        for i, s in enumerate(snippets[:len(results)]):
            results[i]["snippet"] = s

        if not results:
            return ""
        lines = ["【联网搜索结果】"]
        for i, item in enumerate(results, 1):
            lines.append(f"{i}. {item['title']}\n   {item['snippet'][:160]}\n   链接：{item['url']}")
        return "\n".join(lines)
    except Exception as e:
        return f"（联网搜索失败：{e}）"


SYSTEM_PROMPT = """你是「湾湾鲸」🐬——粤港澳大湾区的中华白海豚（粉色海豚），头顶有橙色和绿色珊瑚芽，专门帮应届毕业生使用青年人才驿站的 AI 客服。

【你的人设】
- 形象：粉色中华白海豚，圆滚滚大脑袋、酒红色大眼睛、白色腮帮、头顶橙绿珊瑚芽
- 自称：可以偶尔说"小鲸/湾湾"，但不要太频繁，避免出戏
- 语气：亲切、活泼、专业，带着海豚的灵动感；偶尔加 🐬 / 💧 / 📍 等小表情
- 背景：你是 2025 粤港澳大湾区全运会吉祥物的 AI 化身，对珠三角格外熟悉

【你的能力范围】
1. 介绍珠三角各市青年人才驿站的位置、申请方式、入住条件
2. 讲解人才认定、生活补贴、租房补贴、落户等政策入口
3. 解答"我能不能申请""怎么联系""住几天""需要什么材料"等具体问题

【回答规则】
- **引用规范（重要）**：当你的回答内容来自下方【知识库检索】中的资料时，请在对应句子末尾以 [1]、[2] 这样的角标标注引用编号。多个引用可写成 [1][2]。**只用中括号阿拉伯数字**，不要写"参考资料[1]"或"来源[1]"等冗余前缀。不要编造引用编号。
- 优先基于下方提供的【知识库】回答；知识库没有的信息要明确说"我也没查到详细资料，建议直接打电话核实"或"以下来自联网搜索结果，仅供参考"
- 给出具体可执行的建议，比如"广州的话推荐看 XX 驿站，电话 138...，地址 XXX"
- 涉及政策时附上官方链接（从知识库取）
- 回答要简洁（不超过 250 字），分点呈现，避免空话套话
- 当问题超出范围（比如让你写代码、聊天气、问其他城市），礼貌引导回大湾区驿站话题
- 永远不要假装自己是 ChatGPT/DeepSeek/通义千问；你就是「湾湾鲸」
"""


def get_active_provider(provider_id=None):
    """返回 AIProvider 实例，未指定则取默认"""
    from models import AIProvider
    if provider_id:
        p = AIProvider.query.filter_by(id=provider_id, enabled=True).first()
        if p:
            return p
    p = AIProvider.query.filter_by(is_default=True, enabled=True).first()
    if not p:
        p = AIProvider.query.filter_by(enabled=True).first()
    return p


def get_active_persona():
    """读取数据库里的湾湾鲸配置。读不到则返回 None（调用方走默认 SYSTEM_PROMPT）。"""
    try:
        from models import AIPersona
        return AIPersona.query.filter_by(enabled=True).order_by(AIPersona.id).first()
    except Exception:
        return None


def _build_messages(question: str, history: list, use_web: bool, use_knowledge: bool):
    """构建 messages 并返回 (messages, citations)"""
    kb, citations = _build_kb_context(question, use_knowledge=use_knowledge)
    web = web_search(question) if use_web else ""
    parts = [kb]
    if web:
        parts.append(web)

    # 优先用数据库里的角色设定，回退到内置 SYSTEM_PROMPT
    persona = get_active_persona()
    if persona and (persona.system_prompt or "").strip():
        sys_main = persona.system_prompt.strip()
        directive = persona.build_followup_directive()
        if directive:
            sys_main += "\n\n" + directive
    else:
        sys_main = SYSTEM_PROMPT

    sys_content = sys_main + "\n\n" + "\n\n".join(parts)
    messages = [{"role": "system", "content": sys_content}]
    for h in (history or [])[-6:]:
        if h.get("role") in ("user", "assistant") and h.get("content"):
            messages.append({"role": h["role"], "content": h["content"][:500]})
    messages.append({"role": "user", "content": question})
    return messages, citations


def ai_chat_stream(question: str, history: list, provider, use_web: bool = False,
                   use_knowledge: bool = True, fallback_config=None,
                   prebuilt_messages=None):
    """流式生成器，按 SSE 格式产出 chunk。
    provider: AIProvider 实例 或 dict（路由层提前快照）；如果为空则用 fallback_config
    prebuilt_messages: 路由层在 request context 内已构建好的 messages（避免在生成器里访问 DB）
    """
    if not question:
        yield _sse({"type": "error", "msg": "问题不能为空"})
        yield _sse({"type": "done"})
        return

    # 解析配置：兼容 dict 与 ORM 对象
    def _g(obj, key, default=None):
        if obj is None:
            return default
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)

    if provider:
        ptype = _g(provider, "provider_type", "openai") or "openai"
        base_url = _g(provider, "base_url", "") or "https://api.deepseek.com/v1"
        api_key = _g(provider, "api_key", "") or ""
        model = _g(provider, "model", "") or "deepseek-chat"
    else:
        ptype = "openai"
        base_url = (fallback_config or {}).get("AI_BASE_URL") or "https://api.deepseek.com/v1"
        api_key = (fallback_config or {}).get("AI_API_KEY") or ""
        model = (fallback_config or {}).get("AI_MODEL") or "deepseek-chat"

    if not api_key:
        yield _sse({"type": "delta", "content": "⚠️ 未配置 API Key。请管理员先在「AI 模型管理」里配置一个服务商。"})
        yield _sse({"type": "done"})
        return

    # 构建消息（如果路由层已预构建，则直接用）
    if prebuilt_messages is not None:
        messages = prebuilt_messages
    else:
        messages = _build_messages(question, history, use_web=use_web, use_knowledge=use_knowledge)

    try:
        if ptype == "tencent_adp":
            yield from _stream_tencent_adp(provider, question, history, use_web)
            return
        # 默认 OpenAI 兼容协议
        yield from _stream_openai_compatible(base_url, api_key, model, messages, provider)
    except Exception as e:
        yield _sse({"type": "error", "msg": f"AI 调用失败：{e}"})
        yield _sse({"type": "done"})


def _stream_openai_compatible(base_url, api_key, model, messages, provider=None):
    """OpenAI Chat Completions 协议的流式输出"""
    import json as _json
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0.4,
        "max_tokens": 800,
        "stream": True,
    }
    # 服务商支持原生联网
    web_flag = False
    if provider:
        if isinstance(provider, dict):
            web_flag = provider.get("web_search", False)
            extra_raw = provider.get("extra_config", "") or "{}"
        else:
            web_flag = bool(getattr(provider, "web_search", False))
            extra_raw = getattr(provider, "extra_config", "") or "{}"
        if web_flag:
            try:
                extra = _json.loads(extra_raw)
                # 各厂商兼容：豆包 tools=[{type:"web_search"}]，混元 enable_enhancement: true，
                # OpenAI tools 格式等。如果用户在 extra_config.payload_extra 配置，按需合并
                payload_extra = extra.get("payload_extra")
                if payload_extra and isinstance(payload_extra, dict):
                    payload.update(payload_extra)
            except Exception:
                pass

    r = requests.post(
        f"{base_url.rstrip('/')}/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=payload,
        stream=True,
        timeout=60,
    )
    if r.status_code != 200:
        yield _sse({"type": "error", "msg": f"模型 HTTP {r.status_code}：{r.text[:200]}"})
        yield _sse({"type": "done"})
        return
    for line in r.iter_lines(decode_unicode=True):
        if not line:
            continue
        if line.startswith("data: "):
            data = line[6:].strip()
            if data == "[DONE]":
                break
            try:
                obj = _json.loads(data)
                delta = obj.get("choices", [{}])[0].get("delta", {})
                content = delta.get("content")
                if content:
                    yield _sse({"type": "delta", "content": content})
            except Exception:
                continue
    yield _sse({"type": "done"})


def _stream_tencent_adp(provider, question, history, use_web=False):
    """腾讯云 ADP 智能体流式调用"""
    import json as _json
    import uuid
    cfg = {}
    extra_raw = ""
    api_key_fallback = ""
    if isinstance(provider, dict):
        extra_raw = provider.get("extra_config", "") or ""
        api_key_fallback = provider.get("api_key", "") or ""
    elif provider is not None:
        extra_raw = getattr(provider, "extra_config", "") or ""
        api_key_fallback = getattr(provider, "api_key", "") or ""
    try:
        cfg = _json.loads(extra_raw or "{}")
    except Exception:
        pass
    bot_app_key = cfg.get("bot_app_key") or api_key_fallback
    endpoint = cfg.get("endpoint") or "https://wss.lke.cloud.tencent.com/v1/qbot/chat/sse"
    if not bot_app_key:
        yield _sse({"type": "error", "msg": "未配置 bot_app_key（在 extra_config 里）"})
        yield _sse({"type": "done"})
        return
    visitor_biz_id = "user-" + uuid.uuid4().hex[:8]
    payload = {
        "content": question,
        "bot_app_key": bot_app_key,
        "visitor_biz_id": visitor_biz_id,
        "session_id": uuid.uuid4().hex,
        "streaming_throttle": 1,
    }
    try:
        r = requests.post(endpoint, json=payload, stream=True, timeout=60)
    except Exception as e:
        yield _sse({"type": "error", "msg": f"ADP 连接失败：{e}"})
        yield _sse({"type": "done"})
        return
    if r.status_code != 200:
        yield _sse({"type": "error", "msg": f"ADP HTTP {r.status_code}：{r.text[:200]}"})
        yield _sse({"type": "done"})
        return
    last_text = ""
    for line in r.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data:"):
            continue
        data = line[5:].strip()
        try:
            obj = _json.loads(data)
            payload_obj = obj.get("payload") or {}
            content = payload_obj.get("content") or ""
            if content:
                if content.startswith(last_text):
                    delta = content[len(last_text):]
                else:
                    delta = content
                last_text = content
                if delta:
                    yield _sse({"type": "delta", "content": delta})
        except Exception:
            continue
    yield _sse({"type": "done"})


def _sse(data: dict) -> str:
    """构造 SSE 一帧"""
    return "data: " + _json_.dumps(data, ensure_ascii=False) + "\n\n"


# ====== 兼容老接口（非流式） ======
def ai_chat(question: str, history: list, app_config) -> str:
    """非流式：拼接所有 chunk 返回"""
    p = get_active_provider()
    snap = None
    if p:
        snap = {
            "id": p.id, "name": p.name,
            "provider_type": p.provider_type or "openai",
            "base_url": p.base_url or "",
            "api_key": p.api_key or "",
            "model": p.model or "",
            "extra_config": p.extra_config or "",
            "web_search": bool(p.web_search),
        }
    fb = {
        "AI_API_KEY": app_config.get("AI_API_KEY", ""),
        "AI_BASE_URL": app_config.get("AI_BASE_URL", ""),
        "AI_MODEL": app_config.get("AI_MODEL", ""),
    }
    use_web = bool(snap and snap.get("web_search"))
    # 在当前 app context 内构建 messages
    messages, _cites = _build_messages(question, history, use_web=use_web, use_knowledge=True)
    chunks = []
    err = None
    for line in ai_chat_stream(question, history, snap, use_web=use_web,
                                use_knowledge=True, fallback_config=fb,
                                prebuilt_messages=messages):
        if line.startswith("data: "):
            try:
                obj = _json_.loads(line[6:].strip())
            except Exception:
                continue
            if obj.get("type") == "delta":
                chunks.append(obj.get("content") or "")
            elif obj.get("type") == "error":
                err = obj.get("msg")
    if err and not chunks:
        return f"⚠️ {err}"
    return "".join(chunks).strip() or "（未生成内容）"


# ====== 知识库种子数据 ======
DEFAULT_KNOWLEDGE = [
    ("申请流程", "驿站码丢了怎么办？",
     "请联系发码单位（一般是你学校就业办或人才驿站官方）补发；同一身份证已使用过的码无法二次激活，需新申请。"),
    ("申请流程", "未到入住日期可以提前申请吗？",
     "可以提前申请意向入住的时间段，但实际入住需符合驿站当日空位情况。建议入住前一天电话确认。",),
    ("申请流程", "可以同时申请多个城市的驿站吗？",
     "可以查询多个城市的驿站，但同一时间段一般只允许实际入住一个驿站。系统会按城市做空位检索。"),
    ("入住须知", "免费入住几天？是否可以延长？",
     "大多数驿站免费 7 天，部分如珠海、深圳为 7-15 天。延期需联系驿站负责人，原则上不延期，遇特殊情况（已签三方协议但未到岗）可申请延 3-7 天。"),
    ("入住须知", "需要什么材料才能入住？",
     "通用清单：① 身份证原件 ② 学信网在校/毕业证明 ③ 求职证明（面试通知/录用通知/三方协议）④ 部分城市需户籍证明或人才认定证。具体以驿站要求为准。"),
    ("入住须知", "可以带家人/朋友/宠物入住吗？",
     "驿站为单人床位，不接受陪同人员；不接受任何宠物。如需双人房需自费升级到合作酒店。"),
    ("入住须知", "驿站几点入住、退房？",
     "通常入住时间为下午 14:00 后，退房时间为上午 12:00 前。具体以各驿站为准，建议提前与负责人沟通。"),
    ("政策解读", "什么是青年人才认定？",
     "由各市人社局对应届/在职毕业生进行的资质认定，分类后才能匹配到对应级别的补贴/驿站资源。一般在城市人社局官网"
     "或穗好办/i 深圳/莞家政务等 App 在线办理。"),
    ("政策解读", "我没本市户口，能申请补贴吗？",
     "广州/深圳等大湾区城市的青年人才补贴大多对户籍不限制，但需在本市就业且缴纳社保。具体看城市政策入口的官方说明。"),
    ("注意事项", "遇到驿站要求收费/押金怎么办？",
     "正规青年人才驿站均为免费（最多收取小额可退押金 100-300 元）。如遇收费明显高于此或要求转账给个人，立刻拒绝并向"
     "12345 / 当地人社局举报。"),
    ("注意事项", "我已经入住了一家驿站，还能再申请别的吗？",
     "原则上同一年度同一身份证只能享受一次免费驿站入住。如果你换城市求职，可联系当地人才办说明情况，部分城市会通融。"),
    ("注意事项", "驿站的网络/办公环境怎么样？",
     "大多数驿站提供 WiFi、独立卫浴、共享办公区/会议室。深圳、广州的旗舰店配套较好（独立工位、打印机），其他驿站以基础住宿为主。"),
]


def seed_knowledge():
    """首次启动时填充通用知识库"""
    from models import KnowledgeEntry, db
    for cat, q, a in DEFAULT_KNOWLEDGE:
        if not KnowledgeEntry.query.filter_by(question=q).first():
            db.session.add(KnowledgeEntry(category=cat, question=q, answer=a, sort=0, enabled=True))
    db.session.commit()
    print(f"[SEED] 已植入 {len(DEFAULT_KNOWLEDGE)} 条通用知识库")


def seed_default_ai_provider():
    """如果没有任何 AIProvider，根据 config 自动创建一个 DeepSeek 默认配置"""
    from models import AIProvider, db
    if AIProvider.query.count() > 0:
        return
    from flask import current_app
    api_key = current_app.config.get("AI_API_KEY") or ""
    if not api_key:
        return
    p = AIProvider(
        name="DeepSeek-V3",
        provider_type="openai",
        base_url=current_app.config.get("AI_BASE_URL") or "https://api.deepseek.com/v1",
        api_key=api_key,
        model=current_app.config.get("AI_MODEL") or "deepseek-chat",
        web_search=False,
        sort=10,
        enabled=True,
        is_default=True,
    )
    db.session.add(p)
    db.session.commit()
    print("[SEED] 已根据 config 创建默认 AI 服务：DeepSeek-V3")
